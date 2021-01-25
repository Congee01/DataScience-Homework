import openpyxl
from pyhanlp import *
from xlutils.copy import copy

wb = openpyxl.load_workbook("excel2-bf.xlsx")
sheet = wb["sheet1"]
print("打开成功")

Keyword = ["确诊/v", "病例/n", '联防/vn', '聚集性/nz', '疫情/n', '高风险/n', '联控/n', '联控/v', '境外', "肺炎/nhd", "核酸/n",
           "境外/s", "输入/v", "检疫/n", "检测/vn", "阴性/n", "阳性/n", "密接/nz", "入境/vi", "咽拭子/nz", "隔离/vn", "复学/vi",
           "复课/v", "无症状/nz", "感染者/n", "感染/v", "危重症/n", "医疗队/nis", "疫/ng", "冠状病毒/nhd", "方舱/nz", "火神/nz",
           "雷神/nz", "康复/vn", "口罩/n", "病毒/n"]

max_r = sheet.max_row
max_c = sheet.max_column
col = 1
a = 0

for row in range(1, max_r + 1):
    info = sheet.cell(row, col).value
    if info == None:
        print(info)
        print(row)
        print("这一行是空的")
        continue
    title = HanLP.segment(info)
    print(HanLP.segment(info))
    isMatch = 0
    ival = 0
    for element in title:
        if isMatch == 1:
            break
        for keyword in Keyword:
            if isMatch == 1:
                break
            if str(element) == keyword:
                # 该分词匹配成功
                isMatch = 1
                break
        if isMatch != 1:
            ival += 1

    # 匹配失败 与疫情无关
    if ival == len(title):
        print("匹配失败 与疫情无关")
        print(row)
        sheet.cell(row, col).value = ""
wb.save('excel2-bf.xlsx')
