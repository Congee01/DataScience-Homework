import openpyxl
from pyhanlp import *
HanLP.Config.ShowTermNature = False
CoreStopWordDictionary = JClass("com.hankcs.hanlp.dictionary.stopword.CoreStopWordDictionary")
Pinyin = JClass("com.hankcs.hanlp.dictionary.py.Pinyin")

#打开心态词典
wb1 = openpyxl.load_workbook("Emotion Dictionary-bf.xlsx")
sheet1 = wb1["Sheet1"]
print("打开成功")

max_r = sheet1.max_row
max_c = sheet1.max_column
col = 1
score = [0, 0, 0, 0, 0, 0, 0]
# 下标说明： 0-乐 1-好 2-怒 3-哀 4-惧 5-恶 6-惊
word = []
emotion = []
power = []
# 将心态词典导入列表中
for row in range(2, max_r + 1):
    word_info = sheet1.cell(row, 1).value
    emotion_info = sheet1.cell(row, 5).value
    power_info = sheet1.cell(row, 6).value
    word.append(word_info)
    emotion.append(emotion_info)
    power.append(power_info)
print(word)
print(emotion)
print(power)

# 打开评论文件
wb = openpyxl.load_workbook("excel2-bf.xlsx")
sheet = wb["sheet1"]
max_r = sheet.max_row
max_c = sheet.max_column
print("打开成功")
col = 2

# 央视 注意是从第2列开始读
# 202848~max_r ： 2019.12.8~2020.1.22
# 167787~202847:   2020.1.23~2020.2.7
# 120302~167786: 2020.2.8.2020.3.9
# 1~120301           :2020.3.10~2020.6

# 人民日报 注意是从第一列开始读 打开文件是excel2-bf.xlsx
# 1 ~ 200903     : 2020.3.10~2020.6
# 200904 ~ 281258: 2020.2.8~2020.3.9
# 281259 ~ max_r : 2020.1 ~ 2020.2.7
#
# 测试从500个样本中测试
    # 取得相应评论
    # 将评论分词
#start = 202848
# 央视：
#start = 167787
# start = 120302
# start = 2
# 人民日报：
# start = 1
# start = 200904
start = 281259

#end = max_r + 1
# 央视：
# end = 204848
# end = 167787
# end = 120302
# 人民日报：
# end = 200904
# end = 281259
end = max_r + 1

for row in range(start, end):
        title = sheet.cell(row, 1)
        info = sheet.cell(row, 7).value
        if title is None:
            continue
        print(row)
        str_info = str(info)
        comment = str_info.split("：")[1]
        print(comment)
        segment = HanLP.segment(comment)
        CoreStopWordDictionary.apply(segment)
        print(segment)
        for element in segment:
            # 分词得到的每一项去心态词典相应区域匹配
            # a-60 b-409 c-2152 d-3498 e-4874 f-5002 g-5969 h-7253 j-8923 k-10994
            # l-11561 m-12911 n-13811 o-14163 p-14176 q-14805 r-16334 s-16956 t-19176
            # w-20386 x-21629 y-23259 z-25399
            pinyin_list = HanLP.convertToPinyinList(str(element))
            first_pinyin = str(pinyin_list[0])
            first_character = first_pinyin[0]
            print(first_character)
            if first_character == 'a':
                emotion_row = 60
                end = 408
            elif first_character == 'b':
                emotion_row = 409
                end = 2151
            elif first_character == 'c':
                emotion_row = 2152
                end = 3497
            elif first_character == 'd':
                emotion_row = 3498
                end = 4873
            elif first_character == 'e':
                emotion_row = 4874
                end = 5001
            elif first_character == 'f':
                emotion_row = 5002
                end = 5968
            elif first_character == 'g':
                emotion_row = 5969
                end = 7252
            elif first_character == 'h':
                emotion_row = 7253
                end = 8922
            elif first_character == 'j':
                emotion_row = 8923
                end = 10993
            elif first_character == 'k':
                emotion_row = 10994
                end = 11560
            elif first_character == 'l':
                emotion_row = 11561
                end = 12910
            # a-60 b-409 c-2152 d-3498 e-4874 f-5002 g-5969 h-7253 j-8923 k-10994
            # l-11561 m-12911 n-13811 o-14163 p-14176 q-14805 r-16334 s-16956 t-19176
            # w-20386 x-21629 y-23259 z-25399
            elif first_character == 'm':
                emotion_row = 12911
                end = 13810
            elif first_character == 'n':
                emotion_row = 13811
                end = 14162
            elif first_character == 'o':
                emotion_row = 14163
                end = 14175
            elif first_character == 'p':
                emotion_row = 14176
                end = 14804
            elif first_character == 'q':
                emotion_row = 14805
                end = 16333
            elif first_character == 'r':
                emotion_row = 16334
                end = 16955
            elif first_character == 's':
                emotion_row = 16956
                end = 19175
            elif first_character == 't':
                emotion_row = 19176
                end = 20385
            elif first_character == 'w':
                emotion_row = 20386
                end = 21628
            elif first_character == 'x':
                emotion_row = 21629
                end = 23258
            elif first_character == 'y':
                emotion_row = 23259
                end = 25398
            elif first_character == 'z':
                emotion_row = 25399
                end = 27479
            else:
                emotion_row = 1
                end = 59
            for rowInDict in range(emotion_row, end +1):
                emotion_word = sheet1.cell(rowInDict, 1).value
                if str(emotion_word) == str(element):
                    # 心态词典中有这个词，读取心态种类，强度
                    # 最终得分score[] 下标说明： 0-乐 1-好 2-怒 3-哀 4-惧 5-恶 6-惊
                    emotion_emotion = sheet1.cell(rowInDict, 5).value
                    print(emotion_emotion)
                    emotion_power = sheet1.cell(rowInDict, 6).value
                    print(emotion_power)
                    if emotion_emotion == 'PA' or emotion_emotion == 'PE':
                        score[0] += int(str(emotion_power))
                    elif emotion_emotion == 'PD' or emotion_emotion == 'PH' or emotion_emotion == 'PG' or emotion_emotion ==  'PB' or  emotion_emotion == 'PK':
                        score[1] += int(str(emotion_power))
                    elif emotion_emotion == 'NA':
                        score[2] += int(str(emotion_power))
                    elif emotion_emotion == 'NB' or emotion_emotion == 'NJ'or emotion_emotion == 'NH' or emotion_emotion == 'PF':
                        score[3] += int(str(emotion_power))
                    elif emotion_emotion == 'NI' or emotion_emotion == 'NC' or emotion_emotion == 'NG':
                        score[4] += int(str(emotion_power))
                    elif emotion_emotion == 'NE' or emotion_emotion == 'ND' or emotion_emotion == 'NN' or emotion_emotion == 'NK' or emotion_emotion == 'NL':
                        score[5] += int(str(emotion_power))
                    elif emotion_emotion == 'PC':
                        score[6] += int(str(emotion_power))
                else:
                    continue
print(score)
